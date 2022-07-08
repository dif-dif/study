from collections import Counter
from datetime import date, timedelta
import re
from operator import mod
from openpyxl import Workbook, load_workbook
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

book = load_workbook(filename="C:/Users/Алла/Documents/github/study/exel/original.xlsx")
wb = Workbook()
ws1 = wb.active; ws1.title = 'Доля обращений по сервису'
ws2 = wb.create_sheet('ЮРлицо')
ws3 = wb.create_sheet('Свод за сутки')
ws4 = wb.create_sheet('Доля обращений по НП')
sh = book['01.05']

ws1_last_row = 1; iter = 2
ws1_lastA = 1; ws1_lastD = 1; ws1_lastE = 1
ws2_lastA = 1; ws2_lastD = 1; ws2_lastE = 1
ws3_lastA = 1; ws3_lastD = 1; ws3_lastE = 1

def sh_of_req_service(day):
    global ws1_lastA, ws1_lastD, ws1_lastE, iter

    services = []
    for row in day.iter_rows(min_row=2, min_col=2, max_col=2, max_row=day.max_row, values_only=True):
        services.append(re.sub("[,|(|)|']", "", str(row)))
    count_services = Counter(services)
    print(count_services)
    
    # for service, i in zip(count_services.keys, string.ascii_uppercase[1:]):
    #     ws1[str(i) + '1'].value = service
    
    for row_cells in ws1.iter_rows(min_row=1, max_row=1, min_col=2):
        for cell in row_cells:
            if cell.value in count_services.keys():
                    pass
            else:
                for i in count_services.keys():
                    cell.value = i
            
                
    co = 2
    for col in count_services.values():
        ws1.cell(column=co, row=iter, value=col)
        co += 1


def legal_entity(day):
    pass

def sum_per_day(day):
    # сделать B, C, F
    global ws3_lastA, ws3_lastD, ws3_lastE

    orig_fio = []
    for row in day.iter_rows(min_row=2, min_col=9, max_col=9, max_row=day.max_row, values_only=True):
        orig_fio.append(re.sub("[,|(|)|']", "", str(row)))
    count_appeal = Counter(orig_fio)
    
    def date_convert():
        global ws3_lastA
        title = day.title
        # print(title)
        if title == '2022-04-06':
            for x in range(len(count_appeal)):
                ws3['A'+str(ws3_lastA)].value = '06 Апреля 2022г.'
                ws3_lastA += 1
        elif title == '04.05':
            for x in range(len(count_appeal)):
                ws3['A'+str(ws3_lastA)].value = '04 Мая 2022г.'
                ws3_lastA += 1
        elif title == '05.05':
            for x in range(len(count_appeal)):
                ws3['A'+str(ws3_lastA)].value = '05 Мая 2022г.'
                ws3_lastA += 1
        elif title == '06.05':
            for x in range(len(count_appeal)):
                ws3['A'+str(ws3_lastA)].value = '06 Мая 2022г.'
                ws3_lastA += 1
        else:
            date = re.sub("[\[|\]|'|,]", "", str(title.split('.'))) + ' 2022г.'
            month = {'04': 'Апреля', '05': 'Мая', '06': 'Июня'}
            for key in month.keys():
                date = date.replace(key, str(month[key]))
            for x in range(len(count_appeal)):
                ws3['A'+str(ws3_lastA)].value = date
                ws3_lastA += 1
        return ws3_lastA
    date_convert()
    
    for fio in count_appeal.keys():
        ws3['D'+ str(ws3_lastD)] = fio
        ws3_lastD += 1

    for sum in count_appeal.values():
        ws3['E'+ str(ws3_lastE)] = sum
        ws3_lastE +=1
    
    for cell in range(ws3.max_row):
        ws3['G'+ str(cell+1)].value = 100
        ws3['H'+ str(cell+1)].value = 85
        ws3['I'+ str(cell+1)].value = 100
        ws3['J'+ str(cell+1)].value = 95

    # print(ws3_lastA, ws3_lastD, ws3_lastE)
    return ws3_lastA, ws3_lastD, ws3_lastE

def sh_of_req_np(day):
    pass

def datet():
    d1 = date(2022, 4, 22)
    d2 = date(2022, 6, 10)
    delta = d2-d1
    cou = 1
    for i in range(delta.days+1):
        cou += 1
        ws1['A'+str(cou)].value = d1 + timedelta(i)
        ws2['A'+str(cou)].value = d1 + timedelta(i)
        ws4['A'+str(cou)].value = d1 + timedelta(i)

def services():
    services = []
    for row in book['2022-04-06'].iter_rows(min_row=2, min_col=2, max_col=2, max_row=day.max_row, values_only=True):
        services.append(re.sub("[,|(|)|']", "", str(row)))
    count_services = Counter(services)
    print(count_services)
    
    for row_cells in ws1.iter_rows(min_row=1, max_row=1, min_col=2):
        for cell in row_cells:
            if cell.value in count_services.keys():
                    pass
            else:
                for i in count_services.keys():
                    cell.value = i

for day in enumerate(book.worksheets):
    day = day[1]
    sh_of_req_service(day)
    legal_entity(day)
    sum_per_day(day)
    sh_of_req_np(day)
    iter += 1

datet()

# for row in ws1.values:
#     for value in row:
#         print(value)

wb.save('C:/Users/Алла/Documents/github/study/exel/test.xlsx')
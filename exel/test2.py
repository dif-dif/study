from collections import Counter
from datetime import date, timedelta
import datetime
from html import entities
import re
from operator import mod
import string
from openpyxl import Workbook, load_workbook
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

book = load_workbook(filename="C:/Users/Алла/Documents/github/study/exel/original.xlsx")
wb = Workbook()
ws1 = wb.active; ws1.title = 'Доля обращений по сервису'
ws2 = wb.create_sheet('ЮРлицо')
ws3 = wb.create_sheet('Свод за сутки')
ws4 = wb.create_sheet('Доля обращений по НП')
sh = book['2022-04-06']

dates = {};  services = {}; entitis = {}; objects = {}
column = []
column.extend(list(string.ascii_uppercase[1:]))
for i in list(string.ascii_uppercase):
    column.append('A'+i)

def generation():
    def generate_dict(col:str, dict):
        tmp = []
        for x in enumerate(sh[col]):
            tmp.append(x[1].value)
        tmp = dict.fromkeys(tmp); del tmp[None]
        for i, cell in zip(column, tmp.keys()):
            dict[i] = cell

    for i, j in zip(['B', 'D', 'G'], [services, entitis, objects]):
        generate_dict(i, j)

    def datet():
        d1 = date(2022, 4, 22)
        d2 = date(2022, 6, 10)
        delta = d2-d1
        for i in range(delta.days+1):
            dates[i+2] = str(d1 + timedelta(i))
    datet()

generation()
# print(dates)

class Row:
    def __init__(self, service=None, status=None, entity=None, f_object=None, employee=None, date_create=None, date_closed=None):
        self.service = service
        self.status = status
        self.entity = entity
        self.f_object = f_object
        self.employee = employee
        self.date_create = str(date_create)[:11]
        self.date_closed = str(date_closed)[:11]

rows = []
for row in sh.iter_rows(min_row=2):
    rows.append(Row(row[1].value, row[2].value, row[3].value, row[6].value, row[8].value, row[9].value, row[10].value))

def count_services(name, date_):
    print(len([x for x in rows if x.service == name and x.date_create == date_]))
    print(row[services[name].value].service)
    print(row[dates[date_]].date_create)

    # print(ws[name.keys(name) + dates.keys(date_create)])
def count_organiz(name, date_):
    [x for x in rows if x.entity == name and x.date_create == date_]
    # print(ws[name.keys(name) + dates.keys(date_create)])
def count_object(name, date_):
    [x for x in rows if x.f_object == name and x.date_create == date_]
    # print(ws[name.keys(name) + dates.keys(date_create)])

count_services('Мероприятия', '2022-04-22')


def sh_of_req_service(ws):
    for index_row in dates.keys():
        for index_col in services.keys():
            ws[index_col+index_row].value = count_services(services[index_col], dates[index_row])

def legal_entity():
    pass

def sum_per_day():
    pass

def sh_of_req_np():
    pass



# print(rows)
# print(rows[0].date_create)
# print(1)